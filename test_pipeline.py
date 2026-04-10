"""
test_pipeline.py  —  Test the parser on files you already have.

Usage:
    pip install ixbrlparse openpyxl
    py test_pipeline.py test_files

Outputs:
  - test_results.xlsx        (Excel with 4 tabs, all headers lowercase_underscore)
  - test_results_reports.json (Reports with FULL text — no Excel 32k char limit)
"""

import glob
import json
import os
import re
import sys

# ─── FINANCIAL METRIC MAPPING ───────────────────────────────────────────────
METRIC_MAPPING = {
    "TurnoverRevenue": "revenue", "Turnover": "revenue", "Revenue": "revenue",
    "TurnoverGrossOperatingRevenue": "revenue",
    "GrossProfitLoss": "gross_profit", "GrossProfit": "gross_profit",
    "OperatingProfitLoss": "operating_profit",
    "ProfitLossBeforeTax": "profit_before_tax",
    "ProfitLossOnOrdinaryActivitiesBeforeTax": "profit_before_tax",
    "TaxTaxCreditOnProfitOrLossOnOrdinaryActivities": "tax",
    "ProfitLoss": "profit_after_tax", "ProfitLossForPeriod": "profit_after_tax",
    "ComprehensiveIncomeExpense": "comprehensive_income",
    "CurrentAssets": "current_assets",
    "TradeDebtorsTradeReceivables": "trade_debtors",
    "Debtors": "debtors", "DebtorsTradeOther": "debtors",
    "CashBankOnHand": "cash", "CashBankInHand": "cash",
    "CashCashEquivalents": "cash",
    "TotalAssetsLessCurrentLiabilities": "total_assets_less_current_liabilities",
    "Creditors": "creditors",
    "TradeCreditorsTradePayables": "trade_creditors",
    "NetCurrentAssetsLiabilities": "net_current_assets",
    "NetAssetsLiabilities": "net_assets", "Equity": "equity",
    "ShareholderFunds": "shareholder_funds", "CalledUpShareCapital": "share_capital",
    "AverageNumberEmployeesDuringPeriod": "avg_employees",
    "NetCashFlowsFromUsedInOperatingActivities": "cashflow_operating",
    "NetCashFlowsFromUsedInInvestingActivities": "cashflow_investing",
    "NetCashFlowsFromUsedInFinancingActivities": "cashflow_financing",
    "NetCashGeneratedFromOperations": "net_cash_from_operations",
    "CashCashEquivalentsCashFlowValue": "cash_at_period_end",
}

META_TAGS = {
    "EntityCurrentLegalOrRegisteredName": "company_name",
    "UKCompaniesHouseRegisteredNumber": "company_number_filed",
    "StartDateForPeriodCoveredByReport": "period_start",
    "EndDateForPeriodCoveredByReport": "period_end",
    "BalanceSheetDate": "balance_sheet_date",
    "AccountsType": "accounts_type",
    "AccountingStandardsApplied": "accounting_standard",
    "AccountsStatusAuditedOrUnaudited": "audit_status",
    "EntityDormantTruefalse": "is_dormant",
    "EntityTradingStatus": "trading_status",
    "LegalFormEntity": "legal_form",
    "NameEntityAuditors": "auditor_name",
    "DescriptionPrincipalActivities": "principal_activities",
    "AddressLine1": "address_line_1",
    "AddressLine2": "address_line_2",
    "PrincipalLocation-CityOrTown": "city",
    "CountyRegion": "county",
    "PostalCodeZip": "postcode",
    "NameParentEntity": "parent_entity",
    "NameSubsidiary": "subsidiary_name",
    "DescriptionNatureTransactionsBalancesWithRelatedParties": "related_party_xbrl",
}

DIRECTOR_TAG = "NameEntityOfficer"

# Month lookup for ch_upload column
MONTH_NAMES = {
    "01": "Jan", "02": "Feb", "03": "Mar", "04": "Apr",
    "05": "May", "06": "Jun", "07": "Jul", "08": "Aug",
    "09": "Sep", "10": "Oct", "11": "Nov", "12": "Dec",
}


def strip_ns(raw_name):
    return raw_name.split(":")[-1] if ":" in raw_name else raw_name


def derive_ch_upload(source_file):
    """
    From source_file like 'Prod224_2601_00039175_20250630':
      - '2601' means year=26, month=01 → 'Jan-26'
      - '2512' would mean year=25, month=12 → 'Dec-25'

    Returns e.g. 'Jan-26' or None if can't parse.
    """
    parts = source_file.split("_")
    if len(parts) >= 2:
        batch_code = parts[1]  # e.g. '2601'
        if len(batch_code) == 4:
            yy = batch_code[:2]   # '26'
            mm = batch_code[2:]   # '01'
            month_name = MONTH_NAMES.get(mm)
            if month_name:
                return f"{month_name}-{yy}"
    return None


def derive_date_fields(period_str, instant=None, end=None):
    """
    From the period context, derive:
      - account_closing_date_last_year: the end date as YYYY-MM-DD
      - fiscal_period_new: just the year as int (e.g. 2025)
    """
    closing_date = None
    fiscal_year = None

    if end is not None:
        closing_date = str(end)
    elif instant is not None:
        closing_date = str(instant)
    else:
        if period_str.startswith("as at "):
            closing_date = period_str.replace("as at ", "").strip()
        elif " to " in period_str:
            closing_date = period_str.split(" to ")[-1].strip()

    if closing_date and len(closing_date) >= 4:
        try:
            fiscal_year = int(closing_date[:4])
        except ValueError:
            fiscal_year = None

    return closing_date, fiscal_year


def detect_consolidated(filepath):
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read()
    xbrl = ("ConsolidatedGroupCompanyAccounts" in content or
            ("GroupCompanyDataDimension" in content and "Consolidated" in content) or
            "APConsolidation" in content)
    html = bool(re.search(r'consolidated\s+(financial\s+)?statement', content, re.I) or
                re.search(r'group\s+(strategic|accounts|financial)', content, re.I))
    if xbrl and html:
        return "has_group_but_xbrl_is_parent"
    elif xbrl:
        return "yes"
    elif html:
        return "has_group_but_xbrl_is_parent"
    return "no"


def extract_html_reports(filepath):
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        content = f.read()

    text = re.sub(r'<[^>]+>', '\n', content)
    lines = [l.strip() for l in text.split('\n') if l.strip()]

    section_starts = [
        (re.compile(r'^(GROUP\s+)?STRATEGIC\s+REPORT', re.I), "strategic_report"),
        (re.compile(r'^(REPORT\s+OF\s+THE\s+)?DIRECTORS[\'\u2019]?\s*REPORT', re.I), "directors_report"),
        (re.compile(r'^RELATED\s+PARTY\s+(DISCLOSURES|TRANSACTIONS)', re.I), "related_party_disclosures"),
        (re.compile(r'^(ULTIMATE\s+)?CONTROLL(ING|ED)\s+PARTY', re.I), "ultimate_controlling_party"),
    ]

    stop_patterns = re.compile(
        r'^(REPORT\s+OF\s+THE\s+INDEPENDENT|INDEPENDENT\s+AUDITOR|'
        r'STATEMENT\s+OF\s+COMPREHENSIVE|BALANCE\s+SHEET|'
        r'CONSOLIDATED\s+STATEMENT\s+OF\s+(COMPREHENSIVE|FINANCIAL|CHANGES)|'
        r'NOTES\s+TO\s+THE|CASH\s+FLOW\s+STATEMENT|'
        r'DIRECTORS[\'\u2019]?\s+RESPONSIBILITIES|'
        r'STATEMENT\s+OF\s+CHANGES\s+IN\s+EQUITY|'
        r'COMPANY\s+INFORMATION)', re.I
    )

    skip_patterns = [
        re.compile(r'^Page\s+\d', re.I),
        re.compile(r'^-\s*\d+\s*-$'),
        re.compile(r'^\d+$'),
        re.compile(r'^FOR\s+THE\s+(YEAR|PERIOD)\s+ENDED', re.I),
        re.compile(r'^REGISTERED\s+NUMBER', re.I),
    ]

    continuation_pattern = re.compile(
        r'^(GROUP\s+)?STRATEGIC\s+REPORT\s*\(CONTINUED\)|'
        r'^DIRECTORS[\'\u2019]?\s*REPORT\s*\(CONTINUED\)', re.I
    )

    current_section = None
    sections = {}

    for line in lines:
        new_section = None
        for pattern, section_name in section_starts:
            if pattern.match(line):
                new_section = section_name
                break

        if new_section:
            current_section = new_section
            if current_section not in sections:
                sections[current_section] = []
            continue

        if stop_patterns.match(line):
            current_section = None
            continue

        if current_section and current_section not in ("strategic_report", "directors_report"):
            if re.match(r'^\d{1,2}\.\s*$', line):
                current_section = None
                continue

        if current_section is None:
            continue

        if continuation_pattern.match(line):
            continue
        if any(p.match(line) for p in skip_patterns):
            continue

        sections.setdefault(current_section, []).append(line)

    results = {}
    for section, slines in sections.items():
        text = "\n".join(slines).strip()
        if text and len(text) > 10:
            results[section] = text

    return results


def parse_one_file(filepath):
    from ixbrlparse import IXBRL

    basename = os.path.splitext(os.path.basename(filepath))[0]
    parts = basename.split("_")
    company_number = parts[2] if len(parts) >= 3 else "UNKNOWN"
    filing_date = None
    if len(parts) >= 4 and len(parts[3]) == 8:
        try:
            d = parts[3]
            filing_date = f"{d[:4]}-{d[4:6]}-{d[6:8]}"
        except:
            pass

    ch_upload = derive_ch_upload(basename)
    is_consolidated = detect_consolidated(filepath)
    doc = IXBRL.open(filepath)

    # ── NUMERIC FACTS ──
    financials = []
    seen = set()
    for fact in doc.numeric:
        local = strip_ns(fact.name)
        metric = METRIC_MAPPING.get(local)
        if metric is None:
            continue
        ctx = fact.context
        instant = getattr(ctx, "instant", None)
        start = getattr(ctx, "startdate", None)
        end = getattr(ctx, "enddate", None)
        key = (metric, start, end, instant)
        if key in seen:
            continue
        seen.add(key)
        if fact.value is None:
            continue
        period = ""
        if instant:
            period = f"as at {instant}"
        elif start and end:
            period = f"{start} to {end}"

        if is_consolidated == "has_group_but_xbrl_is_parent":
            data_scope = "parent_only"
        elif is_consolidated == "yes":
            data_scope = "consolidated"
        else:
            data_scope = "single_entity"

        closing_date, fiscal_year = derive_date_fields(period, instant=instant, end=end)

        financials.append({
            "company_number": company_number,
            "metric": metric,
            "value": float(fact.value),
            "period": period,
            "account_closing_date_last_year": closing_date,
            "fiscal_period_new": fiscal_year,
            "is_consolidated": is_consolidated,
            "data_scope": data_scope,
            "filing_date": filing_date,
            "source_file": basename,
            "ch_upload": ch_upload,
        })

    # ── NON-NUMERIC: directors + metadata ──
    company_name = ""
    directors = []
    metadata = []
    text_from_xbrl = []
    directors_seen = set()

    for fact in doc.nonnumeric:
        local = strip_ns(fact.name)
        value = str(fact.value).strip() if fact.value else ""

        if local == "EntityCurrentLegalOrRegisteredName" and value:
            company_name = value

        if local == DIRECTOR_TAG and value and value not in directors_seen:
            directors_seen.add(value)
            directors.append({
                "company_number": company_number,
                "director_name": value,
                "filing_date": filing_date,
                "source_file": basename,
                "ch_upload": ch_upload,
            })

        if local in META_TAGS:
            if local == "DescriptionNatureTransactionsBalancesWithRelatedParties" and value and len(value) > 20:
                text_from_xbrl.append({
                    "company_number": company_number,
                    "company_name": company_name,
                    "section": "related_party_disclosures_xbrl",
                    "text": value,
                    "filing_date": filing_date,
                    "source_file": basename,
                    "ch_upload": ch_upload,
                })
            elif value:
                metadata.append({
                    "company_number": company_number,
                    "field": META_TAGS[local],
                    "value": value,
                    "filing_date": filing_date,
                    "source_file": basename,
                    "ch_upload": ch_upload,
                })

    metadata.append({
        "company_number": company_number,
        "field": "is_consolidated",
        "value": is_consolidated,
        "filing_date": filing_date,
        "source_file": basename,
        "ch_upload": ch_upload,
    })

    # ── HTML REPORT EXTRACTION ──
    html_reports = extract_html_reports(filepath)
    text_sections = list(text_from_xbrl)
    for section_name, text in html_reports.items():
        text_sections.append({
            "company_number": company_number,
            "company_name": company_name,
            "section": section_name,
            "text": text,
            "filing_date": filing_date,
            "source_file": basename,
            "ch_upload": ch_upload,
        })

    return {
        "company_number": company_number,
        "company_name": company_name,
        "filing_date": filing_date,
        "is_consolidated": is_consolidated,
        "financials": financials,
        "directors": directors,
        "text_sections": text_sections,
        "metadata": metadata,
    }


def save_excel(folder, all_financials, all_directors, all_text, all_metadata):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    def write_sheet(ws, headers, rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for row in rows:
            ws.append([v if not isinstance(v, str) or len(v) <= 32767
                        else v[:32760] + "...[TRUNCATED BY EXCEL LIMIT]"
                        for v in row])
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")[:50]) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Sheet 1: Financials
    ws1 = wb.active
    ws1.title = "financials"
    write_sheet(ws1,
        ["company_number", "metric", "value", "period",
         "account_closing_date_last_year", "fiscal_period_new",
         "is_consolidated", "data_scope", "filing_date", "source_file", "ch_upload"],
        [[r["company_number"], r["metric"], r["value"], r["period"],
          r["account_closing_date_last_year"], r["fiscal_period_new"],
          r["is_consolidated"], r["data_scope"], r["filing_date"],
          r["source_file"], r["ch_upload"]]
         for r in all_financials]
    )

    # Sheet 2: Directors
    ws2 = wb.create_sheet("directors")
    write_sheet(ws2,
        ["company_number", "director_name", "filing_date", "source_file", "ch_upload"],
        [[r["company_number"], r["director_name"], r["filing_date"],
          r["source_file"], r["ch_upload"]]
         for r in all_directors]
    )

    # Sheet 3: Reports (text may be truncated by Excel's 32767 limit — use the JSON for full text)
    ws3 = wb.create_sheet("reports")
    write_sheet(ws3,
        ["company_number", "company_name", "section", "text", "filing_date", "source_file", "ch_upload"],
        [[r["company_number"], r["company_name"], r["section"], r["text"],
          r["filing_date"], r["source_file"], r["ch_upload"]]
         for r in all_text]
    )
    ws3.column_dimensions["D"].width = 100
    for row in ws3.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    # Sheet 4: Metadata
    ws4 = wb.create_sheet("metadata")
    write_sheet(ws4,
        ["company_number", "field", "value", "filing_date", "source_file", "ch_upload"],
        [[r["company_number"], r["field"], r["value"], r["filing_date"],
          r["source_file"], r["ch_upload"]]
         for r in all_metadata]
    )

    output_path = os.path.join(folder, "test_results.xlsx")
    wb.save(output_path)
    return output_path


def save_json_reports(folder, all_text):
    """Save reports as JSON — no character limit, full text preserved."""
    output_path = os.path.join(folder, "test_results_reports.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(all_text, f, indent=2, ensure_ascii=False, default=str)
    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: py test_pipeline.py test_files")
        sys.exit(1)

    folder = sys.argv[1]
    if not os.path.isdir(folder):
        print(f"Not a folder: {folder}")
        sys.exit(1)

    files = (
        glob.glob(os.path.join(folder, "*.html"))
        + glob.glob(os.path.join(folder, "*.xhtml"))
        + glob.glob(os.path.join(folder, "*.xml"))
    )
    if not files:
        print(f"No .html / .xhtml / .xml files found in {folder}")
        sys.exit(1)

    test_files = files[:10]
    print(f"\nParsing {len(test_files)} files (out of {len(files)} total)...\n")

    all_financials = []
    all_directors = []
    all_text = []
    all_metadata = []

    for filepath in test_files:
        filename = os.path.basename(filepath)
        print(f"{'─' * 60}")
        print(f"  {filename}")

        try:
            result = parse_one_file(filepath)

            consol_label = {
                "no": "Single entity",
                "yes": "Consolidated (XBRL = group)",
                "has_group_but_xbrl_is_parent": "** GROUP EXISTS but XBRL = parent only **",
            }.get(result["is_consolidated"], result["is_consolidated"])

            print(f"   Company:       {result['company_name']} ({result['company_number']})")
            print(f"   Filing:        {result['filing_date']}")
            print(f"   Consolidated:  {consol_label}")
            print(f"   Financials:    {len(result['financials'])} metrics")
            print(f"   Directors:     {len(result['directors'])} names")
            print(f"   Reports:       {len(result['text_sections'])} sections")

            if result["financials"]:
                print()
                for f in result["financials"]:
                    closing = f['account_closing_date_last_year'] or 'NULL'
                    fy = f['fiscal_period_new'] or 'NULL'
                    print(f"   {f['metric']:40s}  £{f['value']:>14,.0f}   ({f['period']})  closing={closing}  fy={fy}")

            if result["directors"]:
                print(f"\n   Directors: {', '.join(d['director_name'] for d in result['directors'])}")

            if result["text_sections"]:
                for t in result["text_sections"]:
                    print(f"\n   [{t['section']}] {len(t['text'])} chars")
                    preview = t['text'][:200].replace('\n', ' ')
                    print(f"   Preview: {preview}...")

            all_financials.extend(result["financials"])
            all_directors.extend(result["directors"])
            all_text.extend(result["text_sections"])
            all_metadata.extend(result["metadata"])

        except Exception as e:
            print(f"   Error: {e}")
            import traceback
            traceback.print_exc()

        print()

    if all_financials or all_directors or all_text or all_metadata:
        excel_path = save_excel(folder, all_financials, all_directors, all_text, all_metadata)
        json_path = save_json_reports(folder, all_text)
        print(f"{'─' * 60}")
        print(f"Saved Excel:  {excel_path}")
        print(f"Saved JSON:   {json_path}  (full text, no Excel char limit)")
        print(f"   Financials: {len(all_financials)} rows")
        print(f"   Directors:  {len(all_directors)} rows")
        print(f"   Reports:    {len(all_text)} rows")
        print(f"   Metadata:   {len(all_metadata)} rows")

    print(f"\n{'=' * 60}")
    print(f"  Files parsed: {len(test_files)}")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    main()
